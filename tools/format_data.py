#!/usr/bin/env python3
"""
Acoustic Escalation Dashboard — Data Ingester
=============================================
Reads a Zendesk ticket export (.xlsx) and writes data/latest.json,
which the GitHub Pages site fetches on load.

Usage
-----
  python tools/format_data.py path/to/export.xlsx
  python tools/format_data.py path/to/export.xlsx --out data/latest.json
  python tools/format_data.py path/to/export.xlsx --weeks 8 --dry-run

Requirements
------------
  pip install openpyxl
"""

import argparse
import json
import os
import sys
from datetime import date, timedelta

# ── Try importing openpyxl ───────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not found. Run: pip install openpyxl")


# ── Constants ────────────────────────────────────────────────────────────────
REGIONS = ['NA', 'EMEA', 'LATAM', 'APJ']

CATEGORY_MAP = {
    'strained': 'Strained Rel.',
    'product':  'Product',
    'system av':'Product',   # System Availability → Product
    'support':  'Support',
    'techops':  'Support',   # TechOps → Support
    'service':  'Support',   # Service → Support
}

JIRA_BASE = 'https://acoustic-jiraconf.atlassian.net/browse/'


# ── Helpers ──────────────────────────────────────────────────────────────────
def clean(val):
    """Normalise an Excel cell value to a usable string (strips nbsp)."""
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s in ('\xa0', 'None', 'false', 'nan') else s


def parse_date(val):
    """Return a date object from an Excel cell or None."""
    s = clean(val)
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def in_window(d, w_start, w_end):
    """True if d falls in [w_start, w_end)."""
    return d is not None and w_start <= d < w_end


def norm_cat(raw):
    """Map raw Escalation Category to one of the three canonical buckets."""
    if not raw:
        return '—'
    rl = raw.lower()
    for key, bucket in CATEGORY_MAP.items():
        if rl.startswith(key) or key in rl:
            return bucket
    return '—'


def extract_jira(raw):
    """Pull a bare Jira key (e.g. CA-12345) from a URL or return '—'."""
    if not raw:
        return '—'
    if 'browse/' in raw:
        return raw.split('browse/')[-1].split('/')[0].strip()
    import re
    m = re.search(r'[A-Z]+-\d+', raw)
    return m.group(0) if m else (raw or '—')


def compute_rates(e_arr, t_arr):
    """Return list of rounded percentages (1 dp)."""
    return [round(e / t * 100, 1) if t else 0.0 for e, t in zip(e_arr, t_arr)]


def week_windows(n_weeks=8):
    """Return list of (w_start, w_end, label) for the n most recent Mon–Sun weeks."""
    today      = date.today()
    this_monday = today - timedelta(days=today.weekday())
    windows = []
    for i in range(n_weeks, 0, -1):
        ws = this_monday - timedelta(weeks=i)
        we = ws + timedelta(days=7)
        windows.append((ws, we, ws.strftime('%-d %b')))
    return windows


# ── Main processing ──────────────────────────────────────────────────────────
def process(xlsx_path: str, n_weeks: int = 8) -> dict:
    print(f"Loading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    sub_cat_cols = [i for i, h in enumerate(headers) if h == 'Escalation Sub Category']

    # Read all rows into memory
    rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
    print(f"  Rows loaded: {len(rows):,}")

    windows = week_windows(n_weeks)
    w_start_all = windows[0][0]
    w_end_all   = windows[-1][1]
    today       = date.today()

    # ── Per-week denominator / numerator arrays ──────────────────────────
    counts = {
        reg: {'T': [0]*n_weeks, 'E': [0]*n_weeks}
        for reg in REGIONS
    }

    # All escalated tickets in the 8-week window (for enrichment)
    esc_tickets = []

    # Open backlog candidates (any open/pending escalated ticket)
    backlog_raw = []

    for row in rows:
        d = {headers[i]: row[i] for i in range(len(headers))}

        reg = clean(d.get('Region', ''))
        if reg not in REGIONS:
            continue

        org_name = clean(d.get('Ticket organization name', ''))
        if 'demandtec' in org_name.lower():
            continue

        created  = parse_date(d.get('Ticket created - Date'))
        hc_date  = parse_date(d.get('Hot Case Date/Time - Date'))
        ce_date  = parse_date(d.get('Client Escalated Date/Time - Date'))
        hc_flag  = clean(d.get('Hot Case', '')) == 'true'
        ce_flag  = clean(d.get('Client Escalated', '')) == 'true'
        status   = clean(d.get('Ticket status', '')).lower()
        raw_cat  = clean(d.get('Escalation Category', ''))
        csm      = clean(d.get('Planhat CSM', '')) or '—'
        priority = clean(d.get('Ticket priority', '')) or 'Normal'
        jira_raw = clean(d.get('JIRA Ticket Number', ''))
        ticket_id = clean(d.get('Ticket ID', ''))

        sub_cats = [clean(row[i]) for i in sub_cat_cols if clean(row[i])]
        sub = sub_cats[0] if sub_cats else '—'

        jira_key = extract_jira(jira_raw)

        # Esc type
        if hc_flag and ce_flag:
            esc_type = 'Both'
        elif hc_flag:
            esc_type = 'Hot Case'
        elif ce_flag:
            esc_type = 'Client Esc'
        else:
            esc_type = None

        # ── Denominators: ticket created in week ──────────────────────────
        for wi, (ws_d, we_d, lbl) in enumerate(windows):
            if in_window(created, ws_d, we_d):
                counts[reg]['T'][wi] += 1

        # ── Numerators: escalation date (HC or CE) in week ───────────────
        hc_in_any = hc_date and w_start_all <= hc_date < w_end_all
        ce_in_any = ce_date and w_start_all <= ce_date < w_end_all

        if hc_in_any or ce_in_any:
            esc_week_idx = None
            for wi, (ws_d, we_d, _) in enumerate(windows):
                if in_window(hc_date, ws_d, we_d) or in_window(ce_date, ws_d, we_d):
                    esc_week_idx = wi
                    # A + B - C deduplicated count
                    hc_in = in_window(hc_date, ws_d, we_d)
                    ce_in = in_window(ce_date, ws_d, we_d)
                    if hc_in:
                        counts[reg]['E'][wi] = counts[reg]['E'][wi]  # incremented below
                    if ce_in:
                        pass
                    # Unique = 1 per ticket per week
                    counts[reg]['E'][wi] += 1
                    break

            esc_tickets.append({
                'id':       ticket_id,
                'org':      org_name,
                'reg':      reg,
                'esc_type': esc_type or 'Client Esc',
                'cat':      raw_cat or '—',
                'cat_norm': norm_cat(raw_cat),
                'sub':      sub,
                'pri':      priority,
                'st':       clean(d.get('Ticket status', '')) or '—',
                'jira':     jira_key,
                'csm':      csm,
                'week_idx': esc_week_idx,
            })

        # ── Backlog: open/pending escalated tickets ───────────────────────
        if status not in ('solved', 'closed') and (hc_flag or ce_flag):
            if created:
                days_open = (today - created).days
                esc_dates = [dd for dd in [hc_date, ce_date] if dd]
                days_esc  = (today - min(esc_dates)).days if esc_dates else days_open

                backlog_raw.append({
                    'id':       ticket_id,
                    'org':      org_name,
                    'reg':      reg,
                    'esc':      esc_type or 'Client Esc',
                    'cat':      raw_cat or '—',
                    'sub':      sub,
                    'pri':      priority,
                    'st':       clean(d.get('Ticket status', '')) or 'Open',
                    'days_open': days_open,
                    'days_esc':  days_esc,
                    'jira':     jira_key,
                    'csm':      csm,
                })

    # ── Deduplicate escalated counts (A+B−C per week was double-counting) ─
    # Recompute cleanly: one pass, one increment per unique (ticket, week)
    # Reset and recount properly
    for reg in REGIONS:
        counts[reg]['E'] = [0] * n_weeks

    seen = set()
    for t in esc_tickets:
        key = (t['id'], t['week_idx'])
        if key not in seen and t['week_idx'] is not None:
            seen.add(key)
            counts[t['reg']]['E'][t['week_idx']] += 1

    # ── Build output arrays ───────────────────────────────────────────────
    NA_T    = counts['NA']['T'];    NA_E    = counts['NA']['E']
    LATAM_T = counts['LATAM']['T']; LATAM_E = counts['LATAM']['E']
    EMEA_T  = counts['EMEA']['T'];  EMEA_E  = counts['EMEA']['E']
    APJ_T   = counts['APJ']['T'];   APJ_E   = counts['APJ']['E']

    # ── Category breakdown ────────────────────────────────────────────────
    cat_strained = [0] * n_weeks
    cat_product  = [0] * n_weeks
    cat_support  = [0] * n_weeks
    supp_reg     = {'Americas': 0, 'EMEA': 0, 'APJ': 0}
    prod_reg     = {'Americas': 0, 'EMEA': 0, 'APJ': 0}
    type_hc = type_ce = type_both = 0

    for t in esc_tickets:
        if t['week_idx'] is None:
            continue
        nc = t['cat_norm']
        wi = t['week_idx']
        if nc == 'Strained Rel.': cat_strained[wi] += 1
        elif nc == 'Product':     cat_product[wi]  += 1
        elif nc == 'Support':     cat_support[wi]  += 1

        rkey = 'Americas' if t['reg'] in ('NA', 'LATAM') else t['reg']
        if rkey in supp_reg:
            if nc == 'Support': supp_reg[rkey] += 1
            if nc == 'Product': prod_reg[rkey] += 1

        et = t['esc_type']
        if et == 'Hot Case':   type_hc   += 1
        elif et == 'Client Esc': type_ce += 1
        elif et == 'Both':     type_both += 1

    cat_labels = ['Strained Rel.', 'Product', 'Support']
    cat_vals   = [sum(cat_strained), sum(cat_product), sum(cat_support)]
    cat_total  = sum(cat_vals)

    # ── Top 5 accounts ────────────────────────────────────────────────────
    acc = {}
    for t in esc_tickets:
        org = t['org']
        if org not in acc:
            acc[org] = {'n': 0, 'reg': t['reg'], 'csm': t['csm']}
        acc[org]['n'] += 1
    top5 = [
        {'acc': k, 'reg': v['reg'], 'csm': v['csm'], 'n': v['n']}
        for k, v in sorted(acc.items(), key=lambda x: -x[1]['n'])[:5]
    ]

    # ── Last week (W8) tickets ────────────────────────────────────────────
    lw_tickets = [
        {'id': t['id'], 'org': t['org'], 'reg': t['reg'],
         'esc': t['esc_type'], 'cat': t['cat'], 'pri': t['pri'],
         'st':  t['st'],  'jira': t['jira'], 'csm': t['csm']}
        for t in esc_tickets if t['week_idx'] == n_weeks - 1
    ]

    # ── Backlog sort ──────────────────────────────────────────────────────
    backlog = sorted(backlog_raw, key=lambda x: -x['days_esc'])

    # ── Assemble output ───────────────────────────────────────────────────
    weeks_labels = [w[2] for w in windows]
    weeks_starts = [w[0].isoformat() for w in windows]
    weeks_ends   = [w[1].isoformat() for w in windows]

    result = {
        'generated_at': today.isoformat(),
        'weeks':        weeks_labels,
        'week_starts':  weeks_starts,
        'week_ends':    weeks_ends,
        'NA_T':    NA_T,    'LATAM_T': LATAM_T, 'EMEA_T': EMEA_T, 'APJ_T': APJ_T,
        'NA_E':    NA_E,    'LATAM_E': LATAM_E, 'EMEA_E': EMEA_E, 'APJ_E': APJ_E,
        'type_dist':    [type_hc, type_ce, type_both],
        'cat_labels':   cat_labels,
        'cat_vals':     cat_vals,
        'cat_total':    cat_total,
        'cat_strained': cat_strained,
        'cat_product':  cat_product,
        'cat_support':  cat_support,
        'supp_reg':     supp_reg,
        'prod_reg':     prod_reg,
        'res_labels':   cat_labels,
        'res_vals':     [0.0, 0.0, 0.0],  # resolution time: extend when API available
        'top5':         top5,
        'lw_tickets':   lw_tickets,
        'backlog':      backlog,
    }

    # Summary print
    AMER_T = [a+b for a,b in zip(NA_T, LATAM_T)]
    AMER_E = [a+b for a,b in zip(NA_E, LATAM_E)]
    GLOB_T = [a+b+c for a,b,c in zip(AMER_T, EMEA_T, APJ_T)]
    GLOB_E = [a+b+c for a,b,c in zip(AMER_E, EMEA_E, APJ_E)]
    tot_glob_e = sum(GLOB_E)
    tot_glob_t = sum(GLOB_T)
    glob_avg = round(tot_glob_e/tot_glob_t*100, 1) if tot_glob_t else 0

    print(f"\n  8-week window : {weeks_labels[0]} → {weeks_labels[-1]}")
    print(f"  Global avg    : {glob_avg}%  ({tot_glob_e} esc / {tot_glob_t:,} tickets)")
    print(f"  EMEA avg      : {round(sum(EMEA_E)/sum(EMEA_T)*100,1) if sum(EMEA_T) else 0}%")
    print(f"  Americas avg  : {round(sum(AMER_E)/sum(AMER_T)*100,1) if sum(AMER_T) else 0}%")
    print(f"  APJ avg       : {round(sum(APJ_E)/sum(APJ_T)*100,1) if sum(APJ_T) else 0}%")
    print(f"  Open backlog  : {len(backlog)} tickets")
    print(f"  W8 tickets    : {len(lw_tickets)}")
    print(f"  Top account   : {top5[0]['acc']} ({top5[0]['n']} esc)" if top5 else "  No top accounts")

    return result


# ── CLI ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='Convert a Zendesk .xlsx export into data/latest.json for the escalation dashboard.'
    )
    parser.add_argument('xlsx', help='Path to the Zendesk .xlsx export file')
    parser.add_argument('--out', default='data/latest.json',
                        help='Output path (default: data/latest.json)')
    parser.add_argument('--weeks', type=int, default=8,
                        help='Number of complete weeks to include (default: 8)')
    parser.add_argument('--dry-run', action='store_true',
                        help='Process and print summary without writing output')
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        sys.exit(f"ERROR: File not found: {args.xlsx}")

    data = process(args.xlsx, n_weeks=args.weeks)

    if args.dry_run:
        print("\n[dry-run] Output NOT written. Pass without --dry-run to save.")
        return

    out_dir = os.path.dirname(args.out)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)

    size_kb = os.path.getsize(args.out) / 1024
    print(f"\n  Written → {args.out}  ({size_kb:.1f} KB)")
    print("  Next step: git add data/latest.json && git commit -m 'data: "+date.today().isoformat()+"' && git push")


if __name__ == '__main__':
    main()
